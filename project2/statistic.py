import xlrd
from openpyxl import Workbook


# open workbook
wb = xlrd.open_workbook("data.xlsx")

# select sheet data
data_sheet = wb.sheet_by_index(2) # sheet data

output_wb = Workbook()

all_sheet = [output_wb.create_sheet("không"),output_wb.create_sheet("người"),output_wb.create_sheet("trong"),output_wb.create_sheet("có thể"),output_wb.create_sheet("được")]

for i in range(1,38):
    khong_pos = 2
    nguoi_pos = 2
    trong_pos = 2
    co_the_pos = 2
    duoc_pos = 2
    for sheet in all_sheet:
        sheet.cell(row=i, column=1).value = int(data_sheet.cell_value(i,1)) # điền mã số sinh viên
    for j in range(3,18): # duyệt hết các câu
        sentences = data_sheet.cell_value(i,j).split("\n")
        for k in range(2,len(sentences)):
            if(k%2 == 0): # dong chan
                file_path = data_sheet.cell_value(0,j) + "/" + sentences[k-1]
                if("không" in sentences[k]):
                    khong_pos += 1
                    all_sheet[0].cell(row=i, column=khong_pos).value = file_path
                if("người" in sentences[k]):
                    nguoi_pos += 1
                    all_sheet[1].cell(row=i, column=nguoi_pos).value = file_path
                if("trong" in sentences[k]):
                    trong_pos += 1
                    all_sheet[2].cell(row=i, column=trong_pos).value = file_path
                if("có thể" in sentences[k]):
                    co_the_pos += 1
                    all_sheet[3].cell(row=i, column=co_the_pos).value = file_path
                if("được" in sentences[k]):
                    duoc_pos += 1
                    all_sheet[4].cell(row=i, column=duoc_pos).value = file_path
        all_sheet[0].cell(row=i, column=2).value = khong_pos
        all_sheet[1].cell(row=i, column=2).value = nguoi_pos
        all_sheet[2].cell(row=i, column=2).value = trong_pos
        all_sheet[3].cell(row=i, column=2).value = co_the_pos
        all_sheet[4].cell(row=i, column=2).value = duoc_pos
output_wb.save("thongke.xlsx")